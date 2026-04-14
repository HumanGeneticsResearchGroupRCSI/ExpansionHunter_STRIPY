#!/usr/bin/env python3
"""
aggregate_cohort_report.py
==========================
Generate one Excel file per family from per-sample STRipy TSV reports.

File naming : {FamilyID}_{ProbandID}.xlsx
Sheets      : Summary (proband flagged rows) + one sheet per family member
              Sheet order: Summary → Proband → Father → Mother

Proband identified by: phenotype == 2 in PED file
All families generated regardless of whether proband has flagged loci.

Usage
-----
    python3 aggregate_cohort_report.py \\
        --tsv-dir   results/stripy/reports/ \\
        --ped       LH_cnv_wes_hg38_parental.ped \\
        --outdir    results/cohort/

Version : 2.0.0
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# ── Colour scheme ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
PROBAND_FILL  = PatternFill("solid", fgColor="D6E4F0")   # light blue  ← proband rows
PARENT_FILL   = PatternFill("solid", fgColor="FFFFFF")   # white       ← parent rows
PASS_FILL         = PatternFill("solid", fgColor="E2EFDA")   # light green
LOWDEPTH_FILL     = PatternFill("solid", fgColor="FFF2CC")   # light yellow
OUTLIER_HIGH_FILL = PatternFill("solid", fgColor="FCE4D6")   # light orange - HIGH expansion
OUTLIER_LOW_FILL  = PatternFill("solid", fgColor="DDEEFF")   # light blue   - LOW contraction
DENOVO_FILL       = PatternFill("solid", fgColor="F4CCFF")   # light purple - DeNovo
AMBIGUOUS_FILL    = PatternFill("solid", fgColor="E0E0E0")   # medium grey  - AmbiguousNestedCall
NOCALL_FILL       = PatternFill("solid", fgColor="F2F2F2")   # light grey

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# Column indices (0-based) for flagging logic - matches OUTPUT_COLUMNS order
# SampleID=0, FilterStatus=10, CallStatus=11, PopulationOutlier=30,
# PopulationZscore_Direction=32, DeNovo=33
COL_FILTER    = "FilterStatus"
COL_CALL      = "CallStatus"
COL_A1_OUT    = "Allele1_Outlier"
COL_A2_OUT    = "Allele2_Outlier"
COL_A1_DIR    = "Allele1_Direction"
COL_A2_DIR    = "Allele2_Direction"
# DeNovo columns now per-allele: Allele1_DeNovo, Allele2_DeNovo


# ── PED parsing ───────────────────────────────────────────────────────────────

def load_ped(ped_path: str) -> dict:
    """
    Parse PED file.
    Returns dict keyed by family_id:
        {
          "LH0003": {
              "proband":  "LH0003A_S6",
              "father":   "LH0003B_S4",
              "mother":   "LH0003C_S7",
              "members":  ["LH0003A_S6", "LH0003B_S4", "LH0003C_S7"]
          }, ...
        }
    Proband = sample with phenotype == 2.
    Father/mother derived from proband's paternal_id / maternal_id.
    """
    p = Path(ped_path)
    if not p.exists():
        print(f"[ERROR] PED not found: {ped_path}", file=sys.stderr)
        sys.exit(1)

    # First pass - collect all sample rows
    samples = {}   # sample_id → row dict
    with open(p) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            cols = line.split()
            if len(cols) < 6:
                continue
            # Handle both column orders seen in your PED files:
            # Standard: family_id sample_id paternal maternal sex phenotype
            fam_id, sample_id, pat_id, mat_id, sex, pheno = cols[:6]
            samples[sample_id] = {
                "family_id":  fam_id,
                "paternal_id": pat_id,
                "maternal_id": mat_id,
                "sex":         sex,
                "phenotype":   pheno,
            }

    # Second pass - group by family and identify roles
    families = {}
    for sample_id, row in samples.items():
        fam_id = row["family_id"]
        if fam_id not in families:
            families[fam_id] = {
                "proband": None,
                "father":  None,
                "mother":  None,
                "members": [],
            }
        families[fam_id]["members"].append(sample_id)
        if row["phenotype"] == "2":
            families[fam_id]["proband"] = sample_id

    # Proband suffix sets - used as fallback when phenotype != 2
    PROBAND_SUFFIXES = ('A', 'P')   # e.g. LH0003A_S6, CINDI101_EPI101SB_P
    FATHER_SUFFIXES  = ('B', 'D', 'F')
    MOTHER_SUFFIXES  = ('C', 'M')

    def get_sample_suffix(sample_id: str) -> str:
        """
        Extract the role suffix from a sample ID.
        e.g. LH0003A_S6 → 'A'
             CINDI101_EPI101SB_P → 'P'
             LH0014B-RPT_S7 → 'B'
        Looks at the character immediately after the last non-numeric,
        non-separator block before the first underscore or dash suffix.
        """
        import re
        # Match the letter immediately following the family numeric block
        # e.g. LH0003A → A, LH0014B-RPT → B, CINDI101_EPI101SB_P → P
        m = re.search(r'[0-9]([A-Za-z])(?:[_\-]|$)', sample_id)
        if m:
            return m.group(1).upper()
        # Fallback - last uppercase letter before underscore/dash/end
        m = re.search(r'_([A-Z])(?:_|\s*$)', sample_id)
        if m:
            return m.group(1).upper()
        return ''

    # Third pass - derive father/mother from proband's paternal/maternal IDs
    # Also apply suffix-based fallback for families where proband is phenotype=0
    for fam_id, fam in families.items():

        # Fallback: if no proband found by phenotype=2, use suffix
        if fam["proband"] is None:
            for member_id in fam["members"]:
                suffix = get_sample_suffix(member_id)
                if suffix in PROBAND_SUFFIXES:
                    fam["proband"] = member_id
                    break   # take first match

        proband_id = fam["proband"]
        if proband_id and proband_id in samples:
            pat = samples[proband_id]["paternal_id"]
            mat = samples[proband_id]["maternal_id"]
            fam["father"] = pat if pat != "0" else None
            fam["mother"] = mat if mat != "0" else None

        # Additional fallback: if paternal/maternal IDs are 0 in PED but
        # we can identify parents by suffix among family members
        if fam["father"] is None:
            for member_id in fam["members"]:
                if member_id == fam["proband"]:
                    continue
                if get_sample_suffix(member_id) in FATHER_SUFFIXES:
                    fam["father"] = member_id
                    break

        if fam["mother"] is None:
            for member_id in fam["members"]:
                if member_id == fam["proband"]:
                    continue
                if get_sample_suffix(member_id) in MOTHER_SUFFIXES:
                    fam["mother"] = member_id
                    break

    return families


# ── TSV discovery ─────────────────────────────────────────────────────────────

def find_tsv(tsv_dir: Path, sample_id: str) -> Path | None:
    """
    Find TSV file for a sample in tsv_dir.
    Tries: {sample_id}.stripy_report.tsv
    Returns None if not found.
    """
    candidate = tsv_dir / f"{sample_id}.stripy_report.tsv"
    if candidate.exists():
        return candidate
    return None


def parse_tsv(tsv_path: Path) -> tuple:
    """Parse TSV → (headers: list, rows: list[list])."""
    headers = []
    rows    = []
    with open(tsv_path) as f:
        for i, line in enumerate(f):
            cols = line.rstrip('\n').split('\t')
            if i == 0:
                headers = cols
            else:
                rows.append(cols)
    return headers, rows


# ── Row flagging ──────────────────────────────────────────────────────────────

def is_flagged(headers: list, row: list) -> bool:
    """
    Return True if row should appear in Summary sheet.

    Handles both monoallelic ('True') and biallelic ('False/True',
    'True/False', 'True/True') PopulationOutlier values, and both
    monoallelic ('HIGH') and biallelic ('LOW/HIGH', 'HIGH/LOW',
    'HIGH/HIGH') direction values.
    """
    col = {h: i for i, h in enumerate(headers)}

    def get(name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else '.'

    a1_out    = str(get(COL_A1_OUT))
    a2_out    = str(get(COL_A2_OUT))
    a1_dir    = str(get(COL_A1_DIR))
    a2_dir    = str(get(COL_A2_DIR))
    denovo_a1 = get('Allele1_DeNovo')
    denovo_a2 = get('Allele2_DeNovo')
    call      = get(COL_CALL)

    # HIGH outlier - expansion candidate
    is_outlier_high = (
        call == 'CALLED' and (
            (a1_out == 'True' and a1_dir == 'HIGH') or
            (a2_out == 'True' and a2_dir == 'HIGH')
        )
    )

    # LOW outlier - unusual contraction
    is_outlier_low = (
        call == 'CALLED' and (
            (a1_out == 'True' and a1_dir == 'LOW') or
            (a2_out == 'True' and a2_dir == 'LOW')
        )
    )

    is_denovo = (
        denovo_a1 in ('Yes', 'Possible') or
        denovo_a2 in ('Yes', 'Possible')
    )

    return is_outlier_high or is_outlier_low or is_denovo


def row_fill(headers: list, row: list) -> PatternFill | None:
    """Determine row highlight colour by clinical priority."""
    col = {h: i for i, h in enumerate(headers)}

    def get(name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else '.'

    a1_out  = str(get(COL_A1_OUT))
    a2_out  = str(get(COL_A2_OUT))
    a1_dir  = str(get(COL_A1_DIR))
    a2_dir  = str(get(COL_A2_DIR))
    filt    = get(COL_FILTER)
    call    = get(COL_CALL)

    is_outlier_high = (
        (a1_out == 'True' and a1_dir == 'HIGH') or
        (a2_out == 'True' and a2_dir == 'HIGH')
    )
    is_outlier_low = (
        (a1_out == 'True' and a1_dir == 'LOW') or
        (a2_out == 'True' and a2_dir == 'LOW')
    )

    denovo_a1 = get('Allele1_DeNovo')
    denovo_a2 = get('Allele2_DeNovo')
    cq        = str(get('CallQuality'))

    # Priority: DeNovo > HIGH outlier > LOW outlier > Ambiguous > PASS > LowDepth > NoCall
    if denovo_a1 in ('Yes', 'Possible') or denovo_a2 in ('Yes', 'Possible'):
        return DENOVO_FILL
    if is_outlier_high and cq != 'AmbiguousNestedCall':
        return OUTLIER_HIGH_FILL
    if is_outlier_low and cq != 'AmbiguousNestedCall':
        return OUTLIER_LOW_FILL
    if cq == 'AmbiguousNestedCall':
        return AMBIGUOUS_FILL
    if filt == 'PASS' and call == 'CALLED':
        return PASS_FILL
    if filt == 'LowDepth':
        return LOWDEPTH_FILL
    if call == 'NO_CALL':
        return NOCALL_FILL
    return None


# ── Sheet writing ─────────────────────────────────────────────────────────────

COL_WIDTHS = {
    'SampleID': 16, 'Chr': 8, 'Start': 12, 'End': 12,
    'LocationCoordinates': 22, 'Ref': 6, 'Alt': 12,
    'GT': 8, 'LocusCov': 10, 'RepeatLength': 12,
    'FilterStatus': 12, 'CallStatus': 10,
    'Locus': 16, 'LocationRegion': 14, 'RepeatType': 16, 'Motif': 10,
    'Normal Range (Min)': 14, 'Normal Range (Max)': 14,
    'IntermediateRange (Min)': 18, 'IntermediateRange (Max)': 18,
    'PathogenicCutoff': 16,
    'Disease Name': 35, 'Disease OMIM': 14,
    'Inheritance': 22, 'Onset': 14,
    'Stripy_Gene': 14, 'HPO_Gene': 14,
    'HPO Phenotype': 40,
    'Literature:Repeats': 16, 'Literature:DiseaseName': 35,
    'Literature:Inheritance': 22, 'Literature:Range': 14,
    'PopulationOutlier': 16, 'PopulationZscore': 16,
    'PopulationZscore_Direction': 20,
    'DeNovo': 10, 'DeNovo_Direction': 16,
    'DeNovo_Status': 28, 'Caller': 20,
}



def write_legend(wb):
    """
    Write a colour legend to a dedicated 'Legend' sheet in the workbook.
    Kept separate from Summary to avoid phantom rows when pandas reads the file.
    """
    ws = wb.create_sheet(title="Legend")

    legend = [
        ("COLOUR LEGEND",  None,             True,  ""),
        ("Purple",         DENOVO_FILL,       False, "De Novo candidate (Yes or Possible)"),
        ("Orange",         OUTLIER_HIGH_FILL, False, "Population outlier - HIGH direction (expansion above normal range)"),
        ("Blue",           OUTLIER_LOW_FILL,  False, "Population outlier - LOW direction (contraction below normal range)"),
        ("Grey",           AMBIGUOUS_FILL,    False, "Ambiguous call - Nested/Replaced locus with REPCN=0/0 (treat with caution)"),
        ("Green",          PASS_FILL,         False, "Called locus - PASS filter, within normal range"),
        ("Yellow",         LOWDEPTH_FILL,     False, "Called locus - LowDepth filter (low coverage)"),
        ("Light grey",     NOCALL_FILL,       False, "No-call locus - EH could not genotype (REPCN=./.)"),
        ("CallQuality",    None,              True,  ""),
        ("OK",             None,              False, "Called locus - genotype is reliable"),
        ("NoCall",         NOCALL_FILL,       False, "No-call locus - REPCN=./."),
        ("AmbiguousNestedCall", AMBIGUOUS_FILL, False,
         "Nested/Imperfect GCN locus with GT!=0/0 but REPCN=0 - EH detected variant but could not count pathogenic motif"),
        ("DeNovo values",  None,              True,  ""),
        ("Yes",            DENOVO_FILL,       False, "De novo - allele outside normal range, absent in both parents (TRIO)"),
        ("No",             None,              False, "Inherited - at least one parent carries allele in same direction"),
        ("Possible",       DENOVO_FILL,       False, "Possibly de novo - allele outside normal range, absent in available parent (DUO)"),
        (".",              None,              False, "Not assessed - allele within normal range, or insufficient parent data"),
    ]

    for i, (label, fill, bold, desc) in enumerate(legend, start=1):
        swatch            = ws.cell(row=i, column=1, value=label)
        swatch.font       = Font(bold=bold, size=10,
                                 color="FFFFFF" if (bold and fill is None) else "000000")
        swatch.alignment  = Alignment(horizontal='center', vertical='center')
        swatch.border     = THIN_BORDER
        if fill:
            swatch.fill   = fill
        elif bold:
            swatch.fill   = HEADER_FILL
            swatch.font   = Font(bold=True, size=10, color="FFFFFF")

        desc_cell             = ws.cell(row=i, column=2, value=desc)
        desc_cell.alignment   = Alignment(vertical='center', wrap_text=True)
        desc_cell.border      = THIN_BORDER
        if fill:
            desc_cell.fill    = fill
        elif bold:
            desc_cell.fill    = HEADER_FILL
            desc_cell.font    = Font(bold=True, size=10, color="FFFFFF")
        ws.row_dimensions[i].height = 22

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 72


def write_sheet(ws, headers: list, rows: list, freeze: bool = True):
    """Write header + data rows with formatting."""

    # Header row
    for ci, header in enumerate(headers, 1):
        cell              = ws.cell(row=1, column=ci, value=header)
        cell.font         = HEADER_FONT
        cell.fill         = HEADER_FILL
        cell.alignment    = Alignment(horizontal='center',
                                      vertical='center',
                                      wrap_text=True)
        cell.border       = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # Data rows
    for ri, row in enumerate(rows, 2):
        fill = row_fill(headers, row)
        for ci, value in enumerate(row, 1):
            cell           = ws.cell(row=ri, column=ci,
                                     value=value if value != '.' else '')
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border    = THIN_BORDER
            if fill:
                cell.fill  = fill

    # Column widths
    for ci, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = \
            COL_WIDTHS.get(header, 14)

    # Freeze panes + auto-filter
    if freeze:
        ws.freeze_panes = 'B2'
    ws.auto_filter.ref = \
        f"A1:{get_column_letter(len(headers))}1"


def write_summary_sheet(ws, headers: list, proband_rows: list):
    """
    Write Summary sheet - proband flagged rows only.
    Subset of columns most relevant for clinical review.
    """
    summary_cols = [
        "SampleID", "Locus", "Disease Name",
        "GT",
        "Allele1_Repeats", "Allele2_Repeats",
        "Normal Range (Min)", "Normal Range (Max)",
        "PathogenicCutoff",
        "FilterStatus", "CallStatus", "CallQuality",
        "Allele1_Range", "Allele2_Range",
        "Allele1_Zscore", "Allele2_Zscore",
        "Allele1_Outlier", "Allele2_Outlier",
        "Allele1_Direction", "Allele2_Direction",
        "Allele1_DeNovo", "Allele1_DeNovo_Direction",
        "Allele2_DeNovo", "Allele2_DeNovo_Direction",
        "Caller",
    ]

    col_idx = {h: i for i, h in enumerate(headers)}

    # Header
    for ci, col in enumerate(summary_cols, 1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # Flagged rows only
    # Filter flagged rows - exclude empty/blank rows and non-flagged rows
    locus_idx = headers.index('Locus') if 'Locus' in headers else None
    flagged = [
        r for r in proband_rows
        if is_flagged(headers, r)
        and locus_idx is not None
        and locus_idx < len(r)
        and r[locus_idx] not in ('', '.', None)
    ]

    for ri, row in enumerate(flagged, 2):
        fill = row_fill(headers, row)
        for ci, col in enumerate(summary_cols, 1):
            idx   = col_idx.get(col)
            value = row[idx] if idx is not None and idx < len(row) else ''
            cell  = ws.cell(row=ri, column=ci,
                            value=value if value != '.' else '')
            cell.alignment = Alignment(vertical='center')
            cell.border    = THIN_BORDER
            if fill:
                cell.fill  = fill

    # Column widths
    for ci, col in enumerate(summary_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = \
            COL_WIDTHS.get(col, 16)

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = \
        f"A1:{get_column_letter(len(summary_cols))}1"

    return len(flagged)


# ── Per-family Excel writer ───────────────────────────────────────────────────

def write_family_excel(
    family_id:   str,
    proband_id:  str,
    father_id:   str | None,
    mother_id:   str | None,
    tsv_dir:     Path,
    outdir:      Path,
) -> bool:
    """
    Create one Excel file for a family.
    Returns True if file was created, False if proband TSV not found.
    """
    # Must have proband TSV at minimum
    proband_tsv = find_tsv(tsv_dir, proband_id)
    if proband_tsv is None:
        print(f"  [WARN] Proband TSV not found for {proband_id} - skipping family")
        return False

    proband_headers, proband_rows = parse_tsv(proband_tsv)

    # Load parent TSVs if available
    father_data = None
    mother_data = None
    if father_id:
        ft = find_tsv(tsv_dir, father_id)
        if ft:
            father_data = parse_tsv(ft)
        else:
            print(f"  [INFO] Father TSV not found: {father_id}")
    if mother_id:
        mt = find_tsv(tsv_dir, mother_id)
        if mt:
            mother_data = parse_tsv(mt)
        else:
            print(f"  [INFO] Mother TSV not found: {mother_id}")

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Summary (proband flagged rows)
    summary_ws       = wb.active
    summary_ws.title = "Summary"
    n_flagged = write_summary_sheet(
        summary_ws, proband_headers, proband_rows
    )

    # Sheet 2: Proband
    proband_ws       = wb.create_sheet(title=proband_id[:31])
    write_sheet(proband_ws, proband_headers, proband_rows)

    # Sheet 3: Father (if available)
    if father_data:
        fh, fr    = father_data
        father_ws = wb.create_sheet(title=father_id[:31])
        write_sheet(father_ws, fh, fr)

    # Sheet 4: Mother (if available)
    if mother_data:
        mh, mr    = mother_data
        mother_ws = wb.create_sheet(title=mother_id[:31])
        write_sheet(mother_ws, mh, mr)

    # Write legend sheet
    write_legend(wb)

    # Save
    filename = f"{family_id}_{proband_id}.xlsx"
    out_path = outdir / filename
    wb.save(out_path)

    print(f"  Written : {filename}  "
          f"({n_flagged} flagged, "
          f"{'father+' if father_data else ''}"
          f"{'mother' if mother_data else ''} "
          f"{'sheets included' if father_data or mother_data else 'proband only'})")
    return True


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Generate one Excel file per family from STRipy TSV reports"
    )
    p.add_argument('--tsv-dir', required=True,
                   help="Directory containing *.stripy_report.tsv files")
    p.add_argument('--ped',     required=True,
                   help="PED file for family structure")
    p.add_argument('--outdir',  default='.',
                   help="Output directory for Excel files")
    return p.parse_args()


def main():
    args    = parse_args()
    tsv_dir = Path(args.tsv_dir)
    outdir  = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"[INFO] TSV dir : {tsv_dir}")
    print(f"[INFO] PED     : {args.ped}")
    print(f"[INFO] Out dir : {outdir}")

    # Count available TSVs
    tsv_files = list(tsv_dir.glob("*.stripy_report.tsv"))
    print(f"[INFO] TSV files found : {len(tsv_files)}")

    # Load family structure from PED
    families = load_ped(args.ped)
    print(f"[INFO] Families in PED : {len(families)}")
    print()

    # Process each family
    written = skipped = 0
    for family_id, fam in sorted(families.items()):
        proband_id = fam["proband"]
        father_id  = fam["father"]
        mother_id  = fam["mother"]

        if not proband_id:
            print(f"  [WARN] No proband (phenotype=2) found for family {family_id} - skipping")
            skipped += 1
            continue

        print(f"  Family: {family_id:<20} "
              f"Proband: {proband_id:<25} "
              f"Father: {father_id or 'N/A':<25} "
              f"Mother: {mother_id or 'N/A'}")

        ok = write_family_excel(
            family_id  = family_id,
            proband_id = proband_id,
            father_id  = father_id,
            mother_id  = mother_id,
            tsv_dir    = tsv_dir,
            outdir     = outdir,
        )
        if ok:
            written += 1
        else:
            skipped += 1

    # Summary
    print()
    print(f"{'='*60}")
    print(f"  Families processed : {len(families)}")
    print(f"  Excel files written: {written}")
    print(f"  Skipped            : {skipped}  "
          f"(no proband TSV available yet)")
    print(f"  Output directory   : {outdir}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
